# AI Healthcare CE / Certification Auto-Ingest
# Built for: Michigan_Healthcare_CE_Certification_AI_Tracker.xlsx
#
# Purpose:
# - Read certificate PDFs/images from certificates_inbox/
# - Extract key details with AI
# - Map to Michigan EMS, OTA, social work, First Aid/CPR/BLS, and Stop the Bleed tracker fields
# - Append draft rows into the "CE Log" sheet for human review
#
# Human review is required before using any entry for license renewal.

import os
import re
import json
import shutil
from datetime import datetime, date
from calendar import monthrange
from pathlib import Path

from openai import OpenAI
from openpyxl import load_workbook

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import pytesseract
    from PIL import Image
except ImportError:
    pytesseract = None
    Image = None

WORKBOOK_PATH = Path("Michigan_Healthcare_CE_Certification_AI_Tracker.xlsx")
CE_LOG_SHEET = "CE Log"

INPUT_FOLDER = Path("certificates_inbox")
PROCESSED_FOLDER = Path("certificates_processed")
ERROR_FOLDER = Path("certificates_error")

OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

VALID_CREDENTIAL_TYPES = [
    "EMS - EMR/MFR",
    "EMS - EMT",
    "EMS - Specialist/AEMT",
    "EMS - Paramedic",
    "Occupational Therapy Assistant (OTA)",
    "Social Work - LBSW",
    "Social Work - LMSW",
    "Social Work - LLBSW",
    "Social Work - LLMSW",
    "Social Service Technician (SST)",
    "First Aid/CPR/AED",
    "BLS/CPR - Healthcare Provider",
    "Stop the Bleed",
]

VALID_CATEGORIES = [
    "Preparatory",
    "Airway Ventilation",
    "Patient Assessment",
    "Medical",
    "BLS for Healthcare Provider",
    "Trauma",
    "Peds Med Administration Practical",
    "Peds Airway",
    "Peds Assessment",
    "Peds Medical",
    "Peds Trauma",
    "Operations",
    "Emergency Preparedness",
    "Individual Choice",
    "General CE",
    "Pain & Symptom Management",
    "Ethics",
    "Human Trafficking CE",
    "Implicit Bias",
    "Live/Synchronous",
    "First Aid/CPR/AED",
    "BLS/CPR",
    "Stop the Bleed",
    "Certificate Renewal",
    "Other",
]

CATEGORY_ALIASES = {
    "airway": "Airway Ventilation",
    "airway ventilation": "Airway Ventilation",
    "patient assessment": "Patient Assessment",
    "pt assessment": "Patient Assessment",
    "pediatric airway": "Peds Airway",
    "peds airway": "Peds Airway",
    "pediatric assessment": "Peds Assessment",
    "peds assessment": "Peds Assessment",
    "pediatric medical": "Peds Medical",
    "peds medical": "Peds Medical",
    "pediatric trauma": "Peds Trauma",
    "peds trauma": "Peds Trauma",
    "pediatric med administration": "Peds Med Administration Practical",
    "peds med administration": "Peds Med Administration Practical",
    "emergency preparedness": "Emergency Preparedness",
    "operations": "Operations",
    "medical": "Medical",
    "trauma": "Trauma",
    "individual choice": "Individual Choice",
    "elective": "Individual Choice",
    "general": "General CE",
    "general ce": "General CE",
    "pain": "Pain & Symptom Management",
    "pain and symptom management": "Pain & Symptom Management",
    "ethics": "Ethics",
    "human trafficking": "Human Trafficking CE",
    "implicit bias": "Implicit Bias",
    "live": "Live/Synchronous",
    "synchronous": "Live/Synchronous",
    "cpr": "BLS/CPR",
    "bls": "BLS/CPR",
    "first aid": "First Aid/CPR/AED",
    "aed": "First Aid/CPR/AED",
    "stop the bleed": "Stop the Bleed",
    "bleeding control": "Stop the Bleed",
}

HEADER_TO_FIELD = {
    "Log ID": "log_id",
    "Person/Staff": "person_staff",
    "Credential Type": "credential_type",
    "Entry Type": "entry_type",
    "Course/Certificate Title": "title",
    "Provider": "provider",
    "Approval Source": "approval_source",
    "Approval/Activity #": "approval_number",
    "Date Completed": "date_completed",
    "Expiration Date": "expiration_date",
    "CE Hours": "ce_hours",
    "Requirement Category": "requirement_category",
    "Delivery Format": "delivery_format",
    "Live/Synchronous": "live_synchronous",
    "Count for Renewal": "count_for_renewal",
    "Certificate File/URL": "certificate_file_url",
    "AI Extracted": "ai_extracted",
    "AI Confidence": "ai_confidence",
    "Review Status": "review_status",
    "Missing Fields / Notes": "notes",
}

def extract_text(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        if not pdfplumber:
            raise RuntimeError("Missing dependency: pip install pdfplumber")
        with pdfplumber.open(path) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages)
    if ext in {".png", ".jpg", ".jpeg", ".tiff", ".bmp"}:
        if not pytesseract or not Image:
            raise RuntimeError("Missing OCR dependencies: pip install pytesseract pillow")
        return pytesseract.image_to_string(Image.open(path))
    raise ValueError(f"Unsupported file type: {path.suffix}")

def normalize_date(value: str) -> str:
    if not value:
        return ""
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return value

def parse_date(value: str):
    norm = normalize_date(value)
    try:
        return datetime.strptime(norm, "%Y-%m-%d").date()
    except Exception:
        return None

def add_years(d: date, years: int) -> date:
    try:
        return d.replace(year=d.year + years)
    except ValueError:
        return d.replace(year=d.year + years, day=28)

def end_of_month(d: date) -> date:
    return d.replace(day=monthrange(d.year, d.month)[1])

def derive_expiration(row: dict) -> None:
    """Auto-derive CPR/BLS dates when card dates are missing. Always notes for review."""
    if row.get("expiration_date"):
        return
    completed = parse_date(row.get("date_completed", ""))
    if not completed:
        return

    provider = (row.get("provider") or "").lower()
    credential = (row.get("credential_type") or "").lower()
    title = (row.get("title") or "").lower()

    if "american heart" in provider or "aha" in provider or "bls" in credential or "bls" in title:
        # AHA course cards: valid for two years through end of issue month.
        row["expiration_date"] = end_of_month(add_years(completed, 2)).strftime("%Y-%m-%d")
        row["notes"] = (row.get("notes", "") + " | Expiration auto-derived for AHA/BLS pattern; verify card.").strip()
    elif "red cross" in provider or "first aid" in credential or "cpr/aed" in credential:
        # Red Cross CPR/First Aid/AED: 2-year certificate, verify exact card date.
        row["expiration_date"] = add_years(completed, 2).strftime("%Y-%m-%d")
        row["notes"] = (row.get("notes", "") + " | Expiration auto-derived for Red Cross/First Aid/CPR pattern; verify card.").strip()
    elif "stop the bleed" in credential or "stop the bleed" in title:
        # Official Stop the Bleed certificate expiration may depend on employer/program policy.
        row["notes"] = (row.get("notes", "") + " | Stop the Bleed expiration not auto-derived; use employer/program policy.").strip()

def normalize_category(raw: str) -> str:
    if not raw:
        return "Other"
    key = re.sub(r"\s+", " ", raw.strip().lower())
    return CATEGORY_ALIASES.get(key, raw.strip() if raw.strip() in VALID_CATEGORIES else "Other")

def normalize_credential(raw: str) -> str:
    if not raw:
        return ""
    raw_clean = raw.strip()
    if raw_clean in VALID_CREDENTIAL_TYPES:
        return raw_clean
    low = raw_clean.lower()
    if "paramedic" in low:
        return "EMS - Paramedic"
    if "aemt" in low or "specialist" in low:
        return "EMS - Specialist/AEMT"
    if re.search(r"\bemt\b", low):
        return "EMS - EMT"
    if "emr" in low or "mfr" in low:
        return "EMS - EMR/MFR"
    if "occupational therapy assistant" in low or "ota" in low:
        return "Occupational Therapy Assistant (OTA)"
    if "limited master" in low or "llmsw" in low:
        return "Social Work - LLMSW"
    if "limited bachelor" in low or "llbsw" in low:
        return "Social Work - LLBSW"
    if "lmsw" in low or "master" in low:
        return "Social Work - LMSW"
    if "lbsw" in low or "bachelor" in low:
        return "Social Work - LBSW"
    if "social service technician" in low or "sst" in low:
        return "Social Service Technician (SST)"
    if "stop the bleed" in low:
        return "Stop the Bleed"
    if "bls" in low or "healthcare provider cpr" in low:
        return "BLS/CPR - Healthcare Provider"
    if "first aid" in low or "cpr" in low or "aed" in low:
        return "First Aid/CPR/AED"
    return raw_clean

def parse_certificate_with_ai(text: str, filename: str) -> dict:
    schema_hint = {
        "person_staff": "",
        "credential_type": "",
        "entry_type": "CE Course | Certificate Card | Required Training | Other",
        "title": "",
        "provider": "",
        "approval_source": "",
        "approval_number": "",
        "date_completed": "",
        "expiration_date": "",
        "ce_hours": 0,
        "requirement_category": "",
        "delivery_format": "In-person | Live virtual | Hybrid | Online asynchronous | Practical skills | Employer documentation | Other",
        "live_synchronous": "Yes | No | N/A | Unknown",
        "count_for_renewal": "No",
        "ai_confidence": 0.0,
        "review_status": "Needs Review",
        "missing_fields": [],
        "rationale": "",
    }
    system = (
        "You extract continuing education and certification card information for a Michigan healthcare "
        "CE tracker. Return valid JSON only. Do not guess compliance. If unsure, use Needs Review."
    )
    user = f"""
Valid credential types:
{VALID_CREDENTIAL_TYPES}

Valid requirement categories:
{VALID_CATEGORIES}

Extract the certificate fields. Use this JSON shape:
{json.dumps(schema_hint, indent=2)}

Rules:
- Map EMS CE into Michigan categories when certificate text supports it.
- Map OTA and social work courses to General CE, Pain & Symptom Management, Ethics, Human Trafficking CE, Implicit Bias, or Live/Synchronous when supported.
- Map CPR/AED/First Aid/BLS and Stop the Bleed cards as certification/training entries.
- If the certificate lacks hours, completion date, provider, approval source, participant, or category, flag missing_fields and set review_status to Needs Review.
- count_for_renewal should be "No" unless the certificate appears complete enough for human review.

Filename: {filename}

Certificate text:
---
{text}
---
"""
    resp = client.responses.create(
        model=OPENAI_MODEL,
        input=[{"role": "system", "content": system}, {"role": "user", "content": user}],
        # Responses API takes the structured-output spec under `text.format`,
        # not the Chat Completions `response_format` argument.
        text={"format": {"type": "json_object"}},
    )
    data = json.loads(resp.output_text)

    row = {
        "log_id": f"AUTO-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{Path(filename).stem[:20]}",
        "person_staff": data.get("person_staff", ""),
        "credential_type": normalize_credential(data.get("credential_type", "")),
        "entry_type": data.get("entry_type", "Other"),
        "title": data.get("title", ""),
        "provider": data.get("provider", ""),
        "approval_source": data.get("approval_source", ""),
        "approval_number": data.get("approval_number", ""),
        "date_completed": normalize_date(data.get("date_completed", "")),
        "expiration_date": normalize_date(data.get("expiration_date", "")),
        "ce_hours": data.get("ce_hours", 0) or 0,
        "requirement_category": normalize_category(data.get("requirement_category", "")),
        "delivery_format": data.get("delivery_format", "Other"),
        "live_synchronous": data.get("live_synchronous", "Unknown"),
        "count_for_renewal": data.get("count_for_renewal", "No"),
        "certificate_file_url": str(filename),
        "ai_extracted": "Yes",
        "ai_confidence": data.get("ai_confidence", 0),
        "review_status": data.get("review_status", "Needs Review"),
        "notes": data.get("rationale", ""),
    }

    missing = data.get("missing_fields", [])
    if missing:
        row["notes"] = (row["notes"] + " | Missing: " + ", ".join(missing)).strip()

    derive_expiration(row)

    # Human review safety gate.
    if row["ai_confidence"] < 0.80 or row["review_status"] != "Ready":
        row["count_for_renewal"] = "No"
        row["review_status"] = "Needs Review"

    if row["credential_type"] not in VALID_CREDENTIAL_TYPES:
        row["review_status"] = "Needs Review"
        row["notes"] = (row["notes"] + " | Unknown credential type; map manually.").strip()

    return row

def append_to_ce_log(rows: list) -> None:
    if not rows:
        return
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[CE_LOG_SHEET]
    headers = [cell.value for cell in ws[1]]
    for row in rows:
        ws.append([row.get(HEADER_TO_FIELD.get(h, ""), "") for h in headers])
    wb.save(WORKBOOK_PATH)

def process_inbox():
    INPUT_FOLDER.mkdir(exist_ok=True)
    PROCESSED_FOLDER.mkdir(exist_ok=True)
    ERROR_FOLDER.mkdir(exist_ok=True)

    pending = [p for p in sorted(INPUT_FOLDER.iterdir()) if p.is_file()]
    if not pending:
        print("No files found in certificates_inbox.")
        return

    rows = []
    for path in pending:
        try:
            text = extract_text(path)
            if len(text.strip()) < 50:
                raise RuntimeError("Too little text extracted; scan/OCR quality may be poor.")
            row = parse_certificate_with_ai(text, path.name)
            rows.append(row)
            shutil.move(str(path), str(PROCESSED_FOLDER / path.name))
        except Exception as exc:
            print(f"ERROR: {path.name}: {exc}")
            shutil.move(str(path), str(ERROR_FOLDER / path.name))

    append_to_ce_log(rows)
    print(f"Processed {len(rows)} file(s). Review new rows in CE Log.")

if __name__ == "__main__":
    process_inbox()
